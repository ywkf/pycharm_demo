import operator
from dataclasses import replace
import time

# import xlrd
# from xlrd import xldate_as_datetime
import openpyxl


def date_t(dated):
    for i in dated:
        if len(i.strip()) != 0:
            # print('i: ', i)
            return i


month_dict = {
    '1': '一',
    '2': '二',
    '3': '三',
    '4': '四',
    '5': '五',
    '6': '六',
    '7': '七',
    '8': '八',
    '9': '九',
    '10': '十',
    '11': '十一',
    '12': '十二',
}


def excel_tag():

    excel1 = openpyxl.load_workbook('./2022.8.1--2022.8.7 .xlsx', data_only=True)
    # excel1 = xlrd.open_workbook('./2022.8.1--2022.8.7 .xlsx')
    # table1 = excel1.sheet_by_index(6)
    table1 = excel1.worksheets[0]

    rows = table1.max_row
    columns = table1.max_column
    print(rows, columns)

    # 输出特定的行
    for cell in list(table1.rows)[3]:  # 获取第四行的数据
        if cell.value:
            print(cell.value, end=" ")
    print()

    nrow = 0
    ncol = 0
    title = ''
    list_t = []
    list2 = []
    list3 = []
    for row in table1.rows:
        list1 = []

        for cell in row:
            if cell.value:

                v = str(cell.value).strip()
                # print(cell.value, end=" ")
                if len(v) == 1:
                    v = '0' + v
                list1.append(v)
        if row[1].value or row[2].value:
            nrow += 1
            list2.append(list1)
        if len(list1) != 0:
            if operator.contains(list1[-1], '节目单'):
                title = list1[-1]
                print('title: ', title)
                continue
            if operator.contains(list1[-1], '播出日期'):
                date = list1[-1]
                print('date: ', date)
                continue
            if operator.contains(list1[-1], '视频源'):
                list_t = list1
                print('list_t: ', list_t)
                continue
            if len(list1[2]) != 0:
                dict1 = {}
                for i in range(len(list1)):
                    # print('list_t[i]: ', list_t[i])
                    # print('list1[i]: ', list1[i])
                    dict1[list_t[i]] = list1[i]
                # list3.append({'{}'.format(list_t[i]): list1[i]})
                list3.append(dict1)

    for col in table1.columns:
        if col[13].value or col[8].value:
            ncol += 1
    print(nrow, ncol)

    print(list2)
    print(list3)



    # print(table1.row(1)[2].value)
    # print(table1.cell_value(1, 2))
    # print(table1.cell_value(6, 3))
    # print(table1.row_values(3))
    # print(table1.cell_value(3, 2))
    # da = xlrd.xldate.xldate_as_datetime(table1.row_values(3)[4], "%H/%M/%S")
    # da.strftime("%X")
    # da1 = str(da)[-8:]
    # print(da1)


    # dated0 = table1.row_values(1)
    # print('dated0', dated0)
    # dated = date_t(dated0).strip()
    # dated = table1.cell_value(1, 2)
    # print('dated: ', dated)
    # dated1 = dated.split()[1].split('月')
    # y = dated1[0].split('年')[0].strip()
    # m = dated1[0].split('年')[1]
    # d = dated1[1].split('日')[0]
    # w = dated[-1]
    # M = month_dict.get(m) + '月'
    # D = d
    # if len(m) == 1:
    #     m = '0' + m
    # if len(d) == 1:
    #     d = '0' + d

    # print(m, d)

    # list1 = []
    # # rows = table1.nrows
    # nums = 0
    # for i in range(rows):
    #     col = table1.row_values(i)
    #     if i >= 3 and len(col[3].strip()) != 0 and col[5].strip() != '140ST1#':
    #         # dateH = xldate_as_datetime(col[4], 0).strftime('%H')
    #         # dateM = xldate_as_datetime(col[4], 0).strftime('%M')
    #         # dateS = xldate_as_datetime(col[4], 0).strftime('%S')
    #         # print(dateH, ":", dateM, ":", dateS, sep='')
    #         num = ''
    #         num1 = str(col[0]).split('.0')[0]
    #         if len(num1) == 1:
    #             num1 = '0' + num1
    #         # num = m + d + num1
    #         tag = num + col[3].strip()
    #         tag = tag.replace(' ', '')
    #         # time = dateH + ':' + dateM + ':' + dateS
    #
    #         # date = y + '-' + m + '-' + d
    #
    #         # time1 = xlrd.xldate.xldate_as_datetime(col[4], '%X')
    #         # time2 = str(time1)[-8:]
    #
    #
    #         tag_name = tag[6:]
    #         if operator.contains(tag_name, '好剧'):
    #             tag_name = tag_name[:7]
    #         elif operator.contains(tag_name, '剧场'):
    #             tag_name = tag_name[:8]
    #         copy_name = '000000' + tag_name
    #
    #         # if w == '三' and i == 3:
    #         #     tag = num + col[3].strip().replace('                  ', ' ')
    #         #     copy_name = '000000法治测试卡 *（1）'
    #         #     # time = '05:35:00'
    #         # if w == '三' and i == 4:
    #         #     tag = num + col[3].strip().replace('                  ', ' ')
    #         #     copy_name = '000000法治测试卡 *（2）'
    #         # if i == 35:
    #         #     copy_name = '000000法治黄金剧场-03'
    #         # if w == '日' and i == 35:
    #         #     copy_name = '000000法治黄金剧场-3（周日）'
    #
    #         # if w == '三':
    #         #     if i == 3:
    #         #         tag = num + col[3].strip().replace('                  ', ' ')
    #         #         copy_name = '000000法治测试卡 *（1）'
    #         #         # time = '05:35:00'
    #         #     if i == 4:
    #         #         tag = num + col[3].strip().replace('                  ', ' ')
    #         #         copy_name = '000000法治测试卡 *（2）'
    #         #     if i == 30:
    #         #         copy_name = '000000法治黄金剧场-03'
    #         # if w != '三' and i == 35:
    #         #     copy_name = '000000法治黄金剧场-03'
    #         #     if w == '日' and i == 35:
    #         #         copy_name = '000000法治黄金剧场-3（周日）'
    #
    #
    #
    #         time = '00:00:00'
    #         list1.append({'tag': tag, 'time': time, 'copy_name': copy_name, 'month': 1, 'day': 1, 'date': date})
    #         nums += 1
    #
    # print(nums)
    # return list1


excel_tag()
